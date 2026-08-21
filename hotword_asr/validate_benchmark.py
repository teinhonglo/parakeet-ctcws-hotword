from __future__ import annotations

import argparse
import json
import wave
from collections import defaultdict
from pathlib import Path

from .hotwords import compare_vocabularies, load_hotword_list, load_hotword_map


def _report_target(path: Path) -> tuple[str, str] | None:
    if not path.exists():
        return None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        rows = dict(workbook["Summary"].iter_rows(values_only=True))
    except Exception as error:
        print(f"WARNING: cannot read reference report {path}: {error}")
        return None
    recall = next(
        (value for key, value in rows.items() if key and "Recall" in str(key)),
        "unknown",
    )
    mer = next(
        (value for key, value in rows.items() if key and "MER（" in str(key)),
        "unknown",
    )
    return str(recall), str(mer)


def validate(root: Path) -> dict[str, object]:
    root = root.resolve()
    hotword_map = load_hotword_map(root / "hotwords.json")
    all_hotwords = load_hotword_list(root / "all_hotwords.json")
    with (root / "pseudo_transcripts.json").open(encoding="utf-8") as stream:
        pseudo = json.load(stream)
    if not isinstance(pseudo, dict):
        raise ValueError("pseudo_transcripts.json must be an audio_id -> text object")

    audio_ids = set(hotword_map)
    pseudo_ids = {str(key) for key in pseudo}
    wav_paths = {path.stem: path for path in (root / "audio").glob("*.wav")}
    wav_ids = set(wav_paths)
    if audio_ids != pseudo_ids or audio_ids != wav_ids:
        raise ValueError(
            "Benchmark ID mismatch: "
            f"missing-pseudo={sorted(audio_ids - pseudo_ids)}, "
            f"extra-pseudo={sorted(pseudo_ids - audio_ids)}, "
            f"missing-wav={sorted(audio_ids - wav_ids)}, "
            f"extra-wav={sorted(wav_ids - audio_ids)}"
        )

    total_seconds = 0.0
    over_30_seconds = 0
    specs: set[tuple[int, int, int]] = set()
    for path in wav_paths.values():
        with wave.open(str(path), "rb") as audio:
            spec = (audio.getframerate(), audio.getnchannels(), audio.getsampwidth())
            specs.add(spec)
            duration = audio.getnframes() / float(audio.getframerate())
        if duration <= 0:
            raise ValueError(f"Empty WAV file: {path}")
        total_seconds += duration
        over_30_seconds += duration > 30.0
    if specs != {(16000, 1, 2)}:
        raise ValueError(f"Expected 16 kHz mono PCM16 WAV files, got: {sorted(specs)}")

    vocabulary = compare_vocabularies(hotword_map, all_hotwords)
    case_groups: dict[str, list[str]] = defaultdict(list)
    for word in all_hotwords:
        case_groups[word.strip().casefold()].append(word)
    case_duplicates = [values for values in case_groups.values() if len(values) > 1]

    return {
        "audio_count": len(audio_ids),
        "total_seconds": round(total_seconds, 3),
        "over_30_seconds": over_30_seconds,
        "hotword_instances": vocabulary["ground_truth_instances"],
        "hotword_unique_exact": vocabulary["vocabulary_unique"],
        "hotword_unique_casefold": len(case_groups),
        "casefold_duplicates": case_duplicates,
        "missing_from_vocabulary": vocabulary["missing_from_vocabulary"],
        "extra_in_vocabulary": vocabulary["extra_in_vocabulary"],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate the complete hotword benchmark")
    parser.add_argument("--benchmark-dir", type=Path, required=True)
    args = parser.parse_args()
    report = validate(args.benchmark_dir)
    for key, value in report.items():
        print(f"{key:28s}: {value}")
    target = _report_target(args.benchmark_dir / "report.xlsx")
    if target:
        print(f"reference_report_recall     : {target[0]}")
        print(f"reference_report_mer        : {target[1]}")


if __name__ == "__main__":
    main()
