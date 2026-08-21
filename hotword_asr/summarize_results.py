from __future__ import annotations

import argparse
from pathlib import Path

from .io import write_json


REPORTS = (
    ("Parakeet", "Vanilla", "parakeet", "report_vanilla_asr.xlsx"),
    ("Parakeet", "All Hotwords", "parakeet", "report_ctcws_all_hotwords_asr.xlsx"),
    ("Parakeet", "Oracle Hotwords", "parakeet", "report_ctcws_oracle_hotwords_asr.xlsx"),
    ("Nemotron", "Vanilla", "nemotron", "report_vanilla_asr.xlsx"),
    ("Nemotron", "All Hotwords", "nemotron", "report_gpu_pb_all_hotwords_asr.xlsx"),
    ("Nemotron", "Oracle Hotwords", "nemotron", "report_gpu_pb_oracle_hotwords_asr.xlsx"),
    ("Fun-ASR", "Vanilla", "funasr", "report_vanilla_asr.xlsx"),
    ("Fun-ASR", "All Hotwords", "funasr", "report_hotword_all_hotwords_asr.xlsx"),
    ("Fun-ASR", "Oracle Hotwords", "funasr", "report_hotword_oracle_hotwords_asr.xlsx"),
)


def _metrics(path: Path) -> tuple[float, float]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook["Summary"].iter_rows(values_only=True))
    recall = next(value for key, value, *_ in rows if key and "Recall" in str(key))
    mer = next(value for key, value, *_ in rows if key and "MER（" in str(key))
    def as_ratio(value: object) -> float:
        if isinstance(value, (int, float)):
            numeric = float(value)
            return numeric / 100.0 if numeric > 1.0 else numeric
        rendered = str(value).strip()
        return float(rendered.rstrip("%")) / (100.0 if rendered.endswith("%") else 1.0)

    return as_ratio(recall), as_ratio(mer)


def main() -> None:
    parser = argparse.ArgumentParser(description="Summarize all nine benchmark reports")
    parser.add_argument("--parakeet-dir", type=Path, required=True)
    parser.add_argument("--nemotron-dir", type=Path, required=True)
    parser.add_argument("--funasr-dir", type=Path, required=True)
    parser.add_argument("--target-mer", type=float, default=0.15)
    parser.add_argument("--output-json", type=Path)
    args = parser.parse_args()

    print("Model       Condition         Recall     MER       MER delta  <= target")
    print("----------- ----------------- ---------- --------- ---------- ---------")
    missing = []
    roots = {
        "parakeet": args.parakeet_dir,
        "nemotron": args.nemotron_dir,
        "funasr": args.funasr_dir,
    }
    results = []
    for model, condition, root_name, filename in REPORTS:
        path = roots[root_name] / filename
        if not path.exists():
            missing.append(str(path))
            continue
        recall, mer = _metrics(path)
        results.append(
            {
                "model": model,
                "condition": condition,
                "report": str(path.resolve()),
                "hotword_recall": recall,
                "mer": mer,
            }
        )
    if missing:
        raise FileNotFoundError("Missing canonical reports:\n  " + "\n  ".join(missing))

    vanilla = {
        row["model"]: row["mer"] for row in results if row["condition"] == "Vanilla"
    }
    for row in results:
        row["mer_delta_vs_vanilla"] = row["mer"] - vanilla[row["model"]]
        row["meets_target"] = row["mer"] <= args.target_mer
        print(
            f"{row['model']:11s} {row['condition']:17s} "
            f"{row['hotword_recall']:9.2%} {row['mer']:9.2%} "
            f"{row['mer_delta_vs_vanilla']:+9.2%} "
            f"{'YES' if row['meets_target'] else 'NO'}"
        )
    if args.output_json is not None:
        write_json(
            args.output_json,
            {
                "target_mer": args.target_mer,
                "note": (
                    "MER is measured against the benchmark's ASR-generated pseudo "
                    "transcripts, not human ground-truth transcripts."
                ),
                "results": results,
            },
        )


if __name__ == "__main__":
    main()
